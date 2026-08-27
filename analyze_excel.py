import openpyxl

wb = openpyxl.load_workbook("presupuesto_inper.xlsx", read_only=True)
print("Sheet names in workbook:", wb.sheetnames)

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"\n--- Sheet: '{name}' ---")
    row_count = 0
    for i, row in enumerate(sheet.iter_rows(values_only=True)):
        if i < 5:
            print(f"Row {i+1}: {row[:15]}")
        row_count += 1
    print(f"Total rows in '{name}': {row_count}")
