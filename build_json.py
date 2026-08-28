"""
build_json.py — Descarga los archivos Excel directamente de Google Drive
via Service Account y genera public/budget_data.json con datos 100% del sheet.
"""

import json
import sys
import shutil
import os
import io
from datetime import datetime

try:
    import requests
    from google.oauth2.service_account import Credentials
    import openpyxl
except ImportError:
    print("ERROR: Instala dependencias: pip install gspread google-auth requests openpyxl")
    sys.exit(1)

# === CONFIG ===
SERVICE_ACCOUNT_FILE = "service_account.json"
FILE1_ID = "1SW5AkaJ_uTMt5zmeatFT5iIu2Yg7es5W"   # Erogaciones & Pagos Operativos
FILE2_ID = "1SgtkWFKfGcjv-9pWIezz1nkQbRZ8y-DY"   # Presupuesto Oficial AC01 2025

SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly"
]

def log(msg):
    print(msg, flush=True)


def get_credentials():
    try:
        creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        return creds
    except FileNotFoundError:
        log(f"ERROR: No se encontro {SERVICE_ACCOUNT_FILE}")
        sys.exit(1)


def download_excel(creds, file_id, filename):
    """Download an Excel file from Google Drive using service account credentials."""
    log(f"  Descargando {filename} (ID: {file_id})...")

    # Refresh token
    import google.auth.transport.requests
    request = google.auth.transport.requests.Request()
    creds.refresh(request)

    url = f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media"
    headers = {"Authorization": f"Bearer {creds.token}"}
    resp = requests.get(url, headers=headers, stream=True)

    if resp.status_code == 200:
        wb = openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
        log(f"  OK - Tabs disponibles: {wb.sheetnames}")
        return wb
    elif resp.status_code == 403:
        log(f"ERROR 403: El archivo {file_id} no tiene permisos para esta Service Account.")
        log(f"  Comparte el archivo con: visualizador@presupuesto-506721.iam.gserviceaccount.com (Lector)")
        sys.exit(1)
    elif resp.status_code == 404:
        log(f"ERROR 404: Archivo {file_id} no encontrado en Drive.")
        sys.exit(1)
    else:
        log(f"ERROR {resp.status_code}: {resp.text[:300]}")
        sys.exit(1)


def parse_float(val):
    if val is None or val == "" or str(val).strip().lower() in ("n/a", "#n/a", "none"):
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return 0.0


def clean_str(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() in ("none", "n/a", "#n/a", "") else s


def clean_date(val):
    from datetime import datetime as dt
    if isinstance(val, dt):
        return val.strftime("%Y-%m-%d")
    if val is None or val == "":
        return ""
    return str(val).strip()


def process_erogaciones_tab(sheet, cap_code, cap_name, partidas_cat):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    records = []
    for idx, r in enumerate(rows[1:], start=2):
        r = list(r) + [""] * max(0, 28 - len(r))

        ff = clean_str(r[0])
        mes_aplic = None
        try:
            mes_aplic = int(float(r[1])) if r[1] not in (None, "") else None
        except Exception:
            pass

        cta_banc_nombre = clean_str(r[2])
        cuenta_bancaria = clean_str(r[3])
        fecha_pago = clean_date(r[4])
        cheque = clean_str(r[6])
        cr_contra = clean_str(r[8])
        tr = clean_str(r[9])
        tipo_sol = clean_str(r[10])
        contrato = clean_str(r[11])
        factura = clean_str(r[12])
        contrato_interno = clean_str(r[13])
        proveedor = clean_str(r[14])
        concepto = clean_str(r[15])
        imp_parcial = parse_float(r[16])
        imp_total = parse_float(r[17])
        pp = clean_str(r[18])

        ptda_raw = r[19]
        ptda_code = ""
        try:
            ptda_code = str(int(float(ptda_raw))) if ptda_raw not in (None, "") else ""
        except Exception:
            ptda_code = clean_str(ptda_raw)

        act = clean_str(r[20])
        cod_area = clean_str(r[21])
        estatus = clean_str(r[22])
        mes_txt = clean_str(r[23])
        ptda_desc = clean_str(r[26]) if len(r) > 26 else ""
        desc = clean_str(r[27]) if len(r) > 27 else ""

        if not ptda_desc and ptda_code in partidas_cat:
            ptda_desc = f"{ptda_code} - {partidas_cat[ptda_code]}"
        elif not ptda_desc and desc:
            ptda_desc = f"{ptda_code} - {desc}" if ptda_code else desc

        if imp_parcial == 0 and imp_total == 0 and not proveedor:
            continue

        records.append({
            "id": f"{cap_code}-{idx}",
            "capitulo": cap_name,
            "capitulo_code": cap_code,
            "ff": ff,
            "mes_aplic": mes_aplic,
            "cta_banc_nombre": cta_banc_nombre,
            "cuenta_bancaria": cuenta_bancaria,
            "fecha_pago": fecha_pago,
            "cheque": cheque,
            "cr_contra": cr_contra,
            "tr": tr,
            "tipo_sol": tipo_sol,
            "contrato": contrato,
            "factura": factura,
            "contrato_interno": contrato_interno,
            "proveedor": proveedor,
            "concepto": concepto,
            "importe_parcial": imp_parcial,
            "importe_total": imp_total,
            "pp": pp,
            "ptda_code": ptda_code,
            "ptda_desc": ptda_desc,
            "actidad": act,
            "cod_area": cod_area,
            "estatus": estatus,
            "mes_txt": mes_txt,
            "origen": "Dispersion Operativa"
        })

    return records


def process_ac01_tab(sheet, partidas_cat):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    records = []
    for idx, r in enumerate(rows[1:], start=2):
        r = list(r) + [""] * max(0, 20 - len(r))
        if not any(r):
            continue

        ramo = clean_str(r[0])
        unidad = clean_str(r[1])
        pp = clean_str(r[7])

        cap_raw = r[10]
        cap_code = ""
        try:
            cap_code = str(int(float(cap_raw))) if cap_raw not in (None, "") else ""
        except Exception:
            pass

        ptda_raw = r[11]
        ptda_code = ""
        try:
            ptda_code = str(int(float(ptda_raw))) if ptda_raw not in (None, "") else ""
        except Exception:
            pass

        ptda_desc = clean_str(r[12])
        if not ptda_desc and ptda_code in partidas_cat:
            ptda_desc = partidas_cat[ptda_code]

        orig = parse_float(r[17])
        mod = parse_float(r[18])
        dev = parse_float(r[19])

        if not ramo and not pp and orig == 0 and mod == 0 and dev == 0:
            continue

        records.append({
            "id": f"AC01-{idx}",
            "ramo": ramo,
            "unidad": unidad,
            "pp": pp,
            "capitulo_code": cap_code,
            "ptda_code": ptda_code,
            "ptda_desc": ptda_desc,
            "monto_original": orig,
            "monto_modificado": mod,
            "monto_devengado": dev
        })

    return records


def build_ac01_summary(ac01_records):
    totals = {
        "1000": {"orig": 0, "mod": 0, "dev": 0},
        "2000": {"orig": 0, "mod": 0, "dev": 0},
        "3000": {"orig": 0, "mod": 0, "dev": 0},
        "4000": {"orig": 0, "mod": 0, "dev": 0},
        "5000": {"orig": 0, "mod": 0, "dev": 0},
        "total": {"orig": 0, "mod": 0, "dev": 0}
    }
    for item in ac01_records:
        c = item["capitulo_code"]
        if c in totals:
            totals[c]["orig"] += item["monto_original"]
            totals[c]["mod"] += item["monto_modificado"]
            totals[c]["dev"] += item["monto_devengado"]
        totals["total"]["orig"] += item["monto_original"]
        totals["total"]["mod"] += item["monto_modificado"]
        totals["total"]["dev"] += item["monto_devengado"]
    return totals


def main():
    log("Conectando a Google Drive via Service Account...")
    creds = get_credentials()

    # === FILE 1: Erogaciones ===
    log(f"Archivo 1: Erogaciones & Pagos Operativos")
    wb1 = download_excel(creds, FILE1_ID, "presupuesto_inper.xlsx")

    partidas_cat = {}
    if "PARTIDAS" in wb1.sheetnames:
        for r in list(wb1["PARTIDAS"].iter_rows(values_only=True))[1:]:
            if r[0] is not None and r[1] is not None:
                try:
                    code = str(int(float(r[0])))
                    partidas_cat[code] = str(r[1]).strip()
                except Exception:
                    pass
        log(f"  Catalogo partidas: {len(partidas_cat)} entradas")

    log("  Procesando tab '3000'...")
    rec_3000 = process_erogaciones_tab(wb1["3000"], "3000", "Servicios Generales (Cap. 3000)", partidas_cat) if "3000" in wb1.sheetnames else []

    log("  Procesando tab '2000'...")
    rec_2000 = process_erogaciones_tab(wb1["2000"], "2000", "Materiales y Suministros (Cap. 2000)", partidas_cat) if "2000" in wb1.sheetnames else []

    all_records = rec_3000 + rec_2000
    log(f"  Total registros operativos: {len(all_records)}")

    # === FILE 2: AC01 ===
    log(f"Archivo 2: Presupuesto Oficial AC01 2025")
    wb2 = download_excel(creds, FILE2_ID, "sheet_2_inper.xlsx")

    if "Catalogos" in wb2.sheetnames or "Catálogos" in wb2.sheetnames:
        tab_cat = "Catalogos" if "Catalogos" in wb2.sheetnames else "Catálogos"
        for r in list(wb2[tab_cat].iter_rows(values_only=True))[1:]:
            if r[0] is not None and r[1] is not None:
                try:
                    code = str(int(float(r[0])))
                    if code not in partidas_cat:
                        partidas_cat[code] = str(r[1]).strip()
                except Exception:
                    pass

    ac01_tab = None
    for name in wb2.sheetnames:
        if "AC01" in name.upper() or "BASE" in name.upper():
            ac01_tab = name
            break
    if not ac01_tab:
        ac01_tab = wb2.sheetnames[0]

    log(f"  Procesando tab '{ac01_tab}'...")
    ac01_records = process_ac01_tab(wb2[ac01_tab], partidas_cat)
    log(f"  Total registros AC01: {len(ac01_records)}")

    ac01_summary = build_ac01_summary(ac01_records)

    log("\n--- PRESUPUESTO AC01 2025 ---")
    for c, v in ac01_summary.items():
        if v["mod"] > 0 or v["dev"] > 0:
            log(f"  Cap {c}: Orig={v['orig']:,.0f} | Mod={v['mod']:,.0f} | Dev={v['dev']:,.0f}")

    # === BUILD DATASET ===
    dataset = {
        "metadata": {
            "title": "Presupuesto Instituto Nacional de Perinatologia (INPER)",
            "generated_at": datetime.now().isoformat(),
            "total_records": len(all_records),
            "total_ac01_records": len(ac01_records),
            "sheets": [
                {
                    "id": FILE1_ID,
                    "name": "Erogaciones & Pagos Operativos (Cap. 2000 y 3000)",
                    "records_count": len(all_records)
                },
                {
                    "id": FILE2_ID,
                    "name": "Presupuesto Oficial Autorizado AC01 2025 (SHCP / Cuenta Publica)",
                    "records_count": len(ac01_records)
                }
            ],
            "service_account": "visualizador@presupuesto-506721.iam.gserviceaccount.com"
        },
        "ac01_summary": ac01_summary,
        "ac01_records": ac01_records,
        "records": all_records
    }

    os.makedirs("public", exist_ok=True)
    with open("budget_data.json", "w", encoding="utf-8") as f:
        json.dump(dataset, f, ensure_ascii=False, indent=2)
    shutil.copy("budget_data.json", "public/budget_data.json")

    log(f"\nSync exitoso - {len(all_records)} registros operativos + {len(ac01_records)} AC01")
    log(f"Generado: {dataset['metadata']['generated_at']}")


if __name__ == "__main__":
    main()
