import openpyxl

wb = openpyxl.load_workbook("presupuesto_inper.xlsx", data_only=True)

for name in wb.sheetnames:
    sheet = wb[name]
    print(f"================ Sheet: '{name}' ================")
    headers = [cell for cell in next(sheet.iter_rows(values_only=True))]
    print("Headers count:", len(headers))
    for idx, h in enumerate(headers):
        print(f"  Col {idx} ({chr(65+idx) if idx<26 else 'A'+chr(65+idx-26)}): {h}")
    
    rows = list(sheet.iter_rows(values_only=True))
    print(f"\nTotal rows (including header): {len(rows)}")
    print("\nSample Data Row 2:")
    if len(rows) > 1:
        row2 = rows[1]
        for idx, (h, val) in enumerate(zip(headers, row2)):
            print(f"  [{idx}] {h} -> {val} (type: {type(val).__name__})")

