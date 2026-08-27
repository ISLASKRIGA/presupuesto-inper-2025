import openpyxl

wb = openpyxl.load_workbook("sheet_2_inper.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    print(f"\n================ SHEET: '{sheet_name}' (Total Rows: {len(rows)}) ================")
    if not rows:
        continue
    for r_idx, row in enumerate(rows[:10]):
        non_empty = [v for v in row if v is not None]
        if non_empty:
            print(f"Row {r_idx+1} ({len(row)} cols): {row[:20]}")

