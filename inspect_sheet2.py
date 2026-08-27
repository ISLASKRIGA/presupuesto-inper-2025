import requests
import openpyxl

NEW_SHEET_ID = '1SgtkWFKfGcjv-9pWIezz1nkQbRZ8y-DY'

export_url = f"https://docs.google.com/spreadsheets/d/{NEW_SHEET_ID}/export?format=xlsx"
r = requests.get(export_url, allow_redirects=True)

print("Export Status:", r.status_code, "Content length:", len(r.content))

if r.status_code == 200 and len(r.content) > 1000:
    filename = "sheet_2_inper.xlsx"
    with open(filename, "wb") as f:
        f.write(r.content)
    print(f"Saved to {filename}!")
    
    wb = openpyxl.load_workbook(filename, data_only=True)
    print("Sheet names in workbook:", wb.sheetnames)
    
    for name in wb.sheetnames:
        sheet = wb[name]
        rows = list(sheet.iter_rows(values_only=True))
        print(f"\n================ Sheet: '{name}' (Total rows: {len(rows)}) ================")
        if rows:
            print("Headers (Row 1):", rows[0][:15])
            print("\nSample Data Row 2:")
            if len(rows) > 1:
                print("Row 2:", rows[1][:15])
            if len(rows) > 2:
                print("Row 3:", rows[2][:15])
else:
    print("Error fetching sheet:", r.text[:500])
