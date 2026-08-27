import openpyxl
import json

wb = openpyxl.load_workbook("sheet_2_inper.xlsx", data_only=True)

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    print(f"\n================ SHEET: '{sheet_name}' (Total Rows: {len(rows)}) ================")
    for idx, r in enumerate(rows):
        # Print non-empty rows
        if any(r):
            print(f"Row {idx+1}: {r}")

